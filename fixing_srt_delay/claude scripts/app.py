"""
Flask web server for the video transcription tool.

Run with:
    python app.py

Then open http://localhost:5000 in your browser.
"""

import os
import re
import sys
import threading
import time
import uuid
from datetime import datetime
from pathlib import Path

import gdown
import openpyxl
from flask import Flask, render_template, request, jsonify, send_from_directory, abort, redirect, url_for
from dotenv import load_dotenv

import transcribe
import transliterate


# -------------------------------------------------------------------
# Setup
# -------------------------------------------------------------------
load_dotenv()

BASE_DIR = Path(__file__).parent.resolve()
DOWNLOADS_DIR = BASE_DIR / "downloads"
OUTPUTS_DIR = BASE_DIR / "outputs"
UPLOADS_DIR = BASE_DIR / "uploads"   # transient input files for transliteration
DOWNLOADS_DIR.mkdir(exist_ok=True)
OUTPUTS_DIR.mkdir(exist_ok=True)
UPLOADS_DIR.mkdir(exist_ok=True)

SUPPORTED_LANGUAGES = [
    "Hindi", "Telugu", "Bengali", "Tamil", "English",
    "Marathi", "Gujarati", "Kannada", "Malayalam", "Urdu", "Punjabi",
]

MAX_LINKS_PER_BATCH = 3

app = Flask(__name__)
# Cap uploads (used by the transliteration page) at 5 MB. SRT/JSON files for a
# typical 25-minute episode are well under 100 KB; this is a generous ceiling.
app.config["MAX_CONTENT_LENGTH"] = 5 * 1024 * 1024


# -------------------------------------------------------------------
# In-memory batch state
# -------------------------------------------------------------------
# Only one batch runs at a time. This dict holds the current batch's state.
# Fields:
#   id: str
#   status: "running" | "done"
#   source_language: str
#   target_language: str      # "" when translation is off, language name when on
#   prompt: str               # caller-provided prompt override; usually empty
#                             # so the appropriate default variant is picked
#   started_at: iso timestamp
#   videos: list of dicts, each with:
#       index, link, title, context, status, step, elapsed_seconds, started_at,
#       finished_at, error, transcript_file, transcript_xlsx_file,
#       translation_file (always None; translation is now an extra column,
#       not a separate file), log (list of strings)
BATCH_LOCK = threading.Lock()
BATCH = None  # type: dict | None


# -------------------------------------------------------------------
# Helpers
# -------------------------------------------------------------------
GDRIVE_PATTERNS = [
    re.compile(r"drive\.google\.com/file/d/([a-zA-Z0-9_-]+)"),
    re.compile(r"drive\.google\.com/open\?id=([a-zA-Z0-9_-]+)"),
    re.compile(r"drive\.google\.com/uc\?id=([a-zA-Z0-9_-]+)"),
    re.compile(r"[?&]id=([a-zA-Z0-9_-]+)"),
]


def extract_drive_id(link: str) -> str | None:
    for pat in GDRIVE_PATTERNS:
        m = pat.search(link)
        if m:
            return m.group(1)
    return None


def safe_slug(s: str) -> str:
    return re.sub(r"[^a-zA-Z0-9_-]+", "_", s).strip("_") or "file"


def filename_prefix(video: dict) -> str:
    """
    Decide the filename prefix for a video's output files.

    If the user provided a title with at least one alphanumeric character, slug
    it and use that. Otherwise fall back to the original 'videoN' pattern.
    Truncated to 80 chars to keep filenames sane.
    """
    title = (video.get("title") or "").strip()
    fallback = f"video{video['index']}"
    if not title:
        return fallback
    # Reject titles that have no alphanumeric content (slug would be the
    # safe_slug 'file' fallback, which is uglier than videoN).
    if not re.search(r"[a-zA-Z0-9]", title):
        return fallback
    return safe_slug(title)[:80] or fallback


# -------------------------------------------------------------------
# Time range parsing helpers
# -------------------------------------------------------------------
def parse_mmss(value: str) -> int | None:
    """
    Parse an 'mm:ss' string into total seconds. Returns None if the input
    is blank. Raises ValueError with a friendly message if the input is
    present but malformed.

    Accepts: '5:30', '05:30', '0:00', '99:59', and a whole-minute count
    like '5' (treated as 5:00).
    Rejects: 'abc', '5.30', '5:60', '-1:00', '1:2:3', anything non-numeric.
    """
    if value is None:
        return None
    s = value.strip()
    if not s:
        return None

    # Whole-number minutes (no colon): "5" -> 5 minutes = 300 seconds
    if ":" not in s:
        if not s.isdigit():
            raise ValueError(f"'{value}' is not a valid time. Use mm:ss format (e.g. 05:30).")
        return int(s) * 60

    parts = s.split(":")
    if len(parts) != 2:
        raise ValueError(f"'{value}' is not a valid time. Use mm:ss format (e.g. 05:30).")
    mm_str, ss_str = parts
    if not mm_str.isdigit() or not ss_str.isdigit():
        raise ValueError(f"'{value}' is not a valid time. Use mm:ss format (e.g. 05:30).")
    mm = int(mm_str)
    ss = int(ss_str)
    if ss >= 60:
        raise ValueError(f"'{value}' has invalid seconds (must be 0-59). Use mm:ss format.")
    return mm * 60 + ss


def format_mmss(total_seconds: int | None) -> str:
    """Format total seconds back to mm:ss. Used for display in the status table."""
    if total_seconds is None:
        return ""
    m = total_seconds // 60
    s = total_seconds % 60
    return f"{m}:{s:02d}"


def format_segment_label(video: dict) -> str:
    """
    Format a short segment label for the status table. The template prepends
    'Segment: ' so we return just the range part.
    Returns "" if neither start nor end is set.
    """
    start_sec = video.get("segment_start")
    end_sec = video.get("segment_end")
    if start_sec is None and end_sec is None:
        return ""
    if start_sec is not None and end_sec is not None:
        return f"{format_mmss(start_sec)}\u2013{format_mmss(end_sec)}"
    if start_sec is not None:
        return f"from {format_mmss(start_sec)}"
    return f"until {format_mmss(end_sec)}"


def download_drive_video(link: str, target_path: Path, log) -> Path:
    """Download a public Google Drive video to target_path. Returns the actual path used."""
    file_id = extract_drive_id(link)
    if not file_id:
        raise RuntimeError(
            f"Could not extract a Google Drive file ID from the link: {link}"
        )
    url = f"https://drive.google.com/uc?id={file_id}"
    log(f"Downloading from Google Drive (file id: {file_id})...")
    # gdown returns the output path on success, None on failure.
    result = gdown.download(url=url, output=str(target_path), quiet=True)
    if not result:
        raise RuntimeError(
            "gdown failed to download the file. "
            "Make sure the link is shared as 'Anyone with the link'."
        )
    actual = Path(result)
    size_mb = actual.stat().st_size / (1024 * 1024)
    log(f"Download complete: {actual.name} ({size_mb:.1f} MB)")
    return actual


def write_output_file(content: str, filename: str) -> Path:
    """
    Write a transcript output file. CSV files are written with a UTF-8 BOM so
    that Excel opens them correctly without mojibake on Devanagari/Tamil/etc.
    Google Sheets ignores the BOM, so this is harmless for that workflow.
    """
    path = OUTPUTS_DIR / filename
    if filename.lower().endswith(".csv"):
        path.write_text(content, encoding="utf-8-sig")
    else:
        path.write_text(content, encoding="utf-8")
    return path


def write_xlsx_from_csv(csv_text: str, filename: str) -> Path:
    """
    Convert pipe-delimited CSV text into a real .xlsx file.

    Each row from the CSV becomes a row in the spreadsheet, with each pipe-
    separated field placed in its own cell. All cells are stored as plain text
    so Excel/Sheets do not auto-coerce timestamps like '01:23:45:12' into
    dates or strip leading zeros from numbers in dialogue.

    Column widths are auto-sized to fit the longest content per column,
    capped at 80 characters wide so dialogue columns don't sprawl.

    Defensive: rows with fewer or more fields than the header are written
    as-is, never crashes the batch.
    """
    path = OUTPUTS_DIR / filename

    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Transcript"

    # Track max length per column for auto-sizing.
    max_widths: dict[int, int] = {}

    for row_text in csv_text.splitlines():
        if not row_text.strip():
            continue
        fields = row_text.split("|")
        # Append the row. Force every cell to be plain text by pre-setting
        # number_format to '@' (Excel's text format).
        sheet.append(fields)
        current_row = sheet.max_row
        for col_idx, field in enumerate(fields, start=1):
            cell = sheet.cell(row=current_row, column=col_idx)
            cell.number_format = "@"
            # Track width
            field_len = len(field)
            if field_len > max_widths.get(col_idx, 0):
                max_widths[col_idx] = field_len

    # Apply auto-widths, capped at 80 chars wide. Add a small padding of 2.
    for col_idx, width in max_widths.items():
        capped = min(width + 2, 80)
        column_letter = openpyxl.utils.get_column_letter(col_idx)
        sheet.column_dimensions[column_letter].width = capped

    workbook.save(path)
    return path


# -------------------------------------------------------------------
# Downloads folder cleanup
# -------------------------------------------------------------------
def get_downloads_info() -> dict:
    """Return current file count and total size of the downloads/ folder."""
    files = [p for p in DOWNLOADS_DIR.iterdir() if p.is_file()]
    total_bytes = sum(p.stat().st_size for p in files)
    return {
        "file_count": len(files),
        "total_bytes": total_bytes,
        "total_mb": round(total_bytes / (1024 * 1024), 1),
    }


def _files_reserved_for_current_batch() -> set[str]:
    """
    Return the set of download filenames that belong to failed videos in the
    current batch (which might be retried). These should NOT be deleted by
    a safe cleanup.
    """
    reserved = set()
    with BATCH_LOCK:
        if not BATCH:
            return reserved
        batch_id = BATCH["id"]
        for v in BATCH["videos"]:
            # Failed videos may be retried, and retry reuses the existing file.
            # Running videos obviously shouldn't have their file yanked.
            if v["status"] in ("failed", "running"):
                reserved.add(f"batch_{batch_id}_video_{v['index']}.mp4")
    return reserved


def cleanup_downloads(force: bool = False) -> dict:
    """
    Delete files in the downloads/ folder.

    Args:
        force: if True, delete all files. If False (default), skip files that
               belong to failed videos in the current batch (preserves retry capability).

    Returns:
        Dict with 'deleted_count', 'deleted_mb', 'skipped_count', 'errors' (list of strings).
    """
    reserved = set() if force else _files_reserved_for_current_batch()

    deleted_count = 0
    deleted_bytes = 0
    skipped_count = 0
    errors = []

    for path in DOWNLOADS_DIR.iterdir():
        if not path.is_file():
            continue
        if path.name in reserved:
            skipped_count += 1
            continue
        try:
            size = path.stat().st_size
            path.unlink()
            deleted_count += 1
            deleted_bytes += size
        except Exception as e:
            errors.append(f"{path.name}: {e}")

    return {
        "deleted_count": deleted_count,
        "deleted_mb": round(deleted_bytes / (1024 * 1024), 1),
        "skipped_count": skipped_count,
        "errors": errors,
    }


# -------------------------------------------------------------------
# Background worker
# -------------------------------------------------------------------
def run_batch(batch_id: str):
    """Process all videos in the current batch sequentially."""
    global BATCH

    with BATCH_LOCK:
        if not BATCH or BATCH["id"] != batch_id:
            return
        videos = BATCH["videos"]
        source_lang = BATCH["source_language"]
        target_lang = BATCH.get("target_language", "")
        prompt = BATCH["prompt"]

    for video in videos:
        idx = video["index"]

        def log(msg: str, v=video):
            stamp = datetime.now().strftime("%H:%M:%S")
            line = f"[{stamp}] {msg}"
            with BATCH_LOCK:
                v["log"].append(line)
                # Keep log bounded so the UI doesn't get huge.
                if len(v["log"]) > 200:
                    v["log"] = v["log"][-200:]

        with BATCH_LOCK:
            video["status"] = "running"
            video["step"] = "Downloading"
            video["started_at"] = datetime.now().isoformat(timespec="seconds")

        start_time = time.time()
        local_video_path: Path | None = None
        segment_video_path: Path | None = None

        try:
            # Step 1: Download from Drive
            tmp_name = f"batch_{batch_id}_video_{idx}.mp4"
            tmp_path = DOWNLOADS_DIR / tmp_name
            local_video_path = download_drive_video(video["link"], tmp_path, log)

            # Step 2: Transcribe (with optional segment range passed in the prompt)
            with BATCH_LOCK:
                video["step"] = "Transcribing"
            start_sec = video.get("segment_start")
            end_sec = video.get("segment_end")
            transcript = transcribe.transcribe_video(
                str(local_video_path),
                source_language=source_lang,
                video_title=video.get("title", ""),
                video_context=video.get("context", ""),
                target_language=target_lang,
                start_time=format_mmss(start_sec),
                end_time=format_mmss(end_sec),
                prompt=prompt,
                log_callback=log,
            )
            transcript_filename = f"{filename_prefix(video)}_{safe_slug(source_lang)}_transcript.csv"
            write_output_file(transcript, transcript_filename)
            with BATCH_LOCK:
                video["transcript_file"] = transcript_filename
            log(f"Transcript saved: {transcript_filename} ({len(transcript)} chars)")

            # Also write an .xlsx version so users can open it directly in
            # Google Sheets without specifying a delimiter.
            try:
                xlsx_filename = f"{filename_prefix(video)}_{safe_slug(source_lang)}_transcript.xlsx"
                write_xlsx_from_csv(transcript, xlsx_filename)
                with BATCH_LOCK:
                    video["transcript_xlsx_file"] = xlsx_filename
                log(f"XLSX saved: {xlsx_filename}")
            except Exception as e:
                # XLSX generation is a convenience; never fail the job over it.
                log(f"Warning: could not generate XLSX: {e}")

            # Sanity check: flag suspiciously short transcripts. The CSV format
            # is much longer than the previous plain-text format, so the
            # threshold is higher here than it was before.
            if len(transcript) < 2000:
                log(
                    "WARNING: transcript is under 2000 characters, which is short for "
                    "a 20-25 minute episode. You may want to retry."
                )

            # NOTE: Translation step intentionally removed. translate_text() in
            # transcribe.py is preserved as dead code for future re-enablement.

            # Success: clean up the local video files (full download + segment).
            try:
                if local_video_path and local_video_path.exists():
                    local_video_path.unlink()
                    log("Cleaned up local video file.")
            except Exception as e:
                log(f"Warning: could not delete local video file: {e}")
            try:
                if segment_video_path and segment_video_path.exists():
                    segment_video_path.unlink()
                    log("Cleaned up local segment file.")
            except Exception as e:
                log(f"Warning: could not delete local segment file: {e}")

            with BATCH_LOCK:
                video["status"] = "done"
                video["step"] = "Done"
                video["finished_at"] = datetime.now().isoformat(timespec="seconds")
                video["elapsed_seconds"] = int(time.time() - start_time)

        except Exception as e:
            log(f"ERROR: {e}")
            with BATCH_LOCK:
                video["status"] = "failed"
                video["step"] = "Failed"
                video["error"] = str(e)
                video["finished_at"] = datetime.now().isoformat(timespec="seconds")
                video["elapsed_seconds"] = int(time.time() - start_time)
            # We keep the downloaded video on failure so the user can retry
            # without re-downloading. It will be overwritten on retry.
            # Continue with the next video.

    with BATCH_LOCK:
        if BATCH and BATCH["id"] == batch_id:
            BATCH["status"] = "done"
            BATCH["finished_at"] = datetime.now().isoformat(timespec="seconds")

    # Auto-cleanup: any download files still sitting around that belong to
    # DONE videos (not failed ones that might be retried) should go.
    # The per-video success path already deletes its own file, so this is
    # a belt-and-braces sweep for anything the success path missed (e.g. crash).
    try:
        cleanup_downloads(force=False)
    except Exception:
        pass  # Cleanup is best-effort; never let it fail the batch.


def retry_single_video(video: dict, source_lang: str, target_lang: str, prompt: str, batch_id: str):
    """Retry a single failed video. Runs in its own thread."""
    idx = video["index"]

    def log(msg: str, v=video):
        stamp = datetime.now().strftime("%H:%M:%S")
        line = f"[{stamp}] {msg}"
        with BATCH_LOCK:
            v["log"].append(line)
            if len(v["log"]) > 200:
                v["log"] = v["log"][-200:]

    with BATCH_LOCK:
        video["status"] = "running"
        video["step"] = "Downloading"
        video["error"] = None
        video["started_at"] = datetime.now().isoformat(timespec="seconds")
        video["finished_at"] = None
        video["elapsed_seconds"] = 0
        video["transcript_file"] = None
        video["transcript_xlsx_file"] = None
        video["translation_file"] = None  # always None now; preserved for compatibility

    start_time = time.time()
    local_video_path: Path | None = None
    segment_video_path: Path | None = None

    try:
        tmp_name = f"batch_{batch_id}_video_{idx}.mp4"
        tmp_path = DOWNLOADS_DIR / tmp_name

        # If the downloaded file from the previous failed attempt still exists
        # and looks healthy, reuse it.
        if tmp_path.exists() and tmp_path.stat().st_size > 0:
            log(f"Reusing previously downloaded file: {tmp_path.name}")
            local_video_path = tmp_path
        else:
            local_video_path = download_drive_video(video["link"], tmp_path, log)

        # If a time range was requested, it will be passed in the prompt.
        start_sec = video.get("segment_start")
        end_sec = video.get("segment_end")

        with BATCH_LOCK:
            video["step"] = "Transcribing"
        transcript = transcribe.transcribe_video(
            str(local_video_path),
            source_language=source_lang,
            video_title=video.get("title", ""),
            video_context=video.get("context", ""),
            target_language=target_lang,
            start_time=format_mmss(start_sec),
            end_time=format_mmss(end_sec),
            prompt=prompt,
            log_callback=log,
        )
        transcript_filename = f"{filename_prefix(video)}_{safe_slug(source_lang)}_transcript.csv"
        write_output_file(transcript, transcript_filename)
        with BATCH_LOCK:
            video["transcript_file"] = transcript_filename
        log(f"Transcript saved: {transcript_filename} ({len(transcript)} chars)")

        try:
            xlsx_filename = f"{filename_prefix(video)}_{safe_slug(source_lang)}_transcript.xlsx"
            write_xlsx_from_csv(transcript, xlsx_filename)
            with BATCH_LOCK:
                video["transcript_xlsx_file"] = xlsx_filename
            log(f"XLSX saved: {xlsx_filename}")
        except Exception as e:
            log(f"Warning: could not generate XLSX: {e}")

        # NOTE: Translation step intentionally removed.

        try:
            if local_video_path and local_video_path.exists():
                local_video_path.unlink()
        except Exception:
            pass
        try:
            if segment_video_path and segment_video_path.exists():
                segment_video_path.unlink()
        except Exception:
            pass

        with BATCH_LOCK:
            video["status"] = "done"
            video["step"] = "Done"
            video["finished_at"] = datetime.now().isoformat(timespec="seconds")
            video["elapsed_seconds"] = int(time.time() - start_time)

    except Exception as e:
        log(f"ERROR: {e}")
        with BATCH_LOCK:
            video["status"] = "failed"
            video["step"] = "Failed"
            video["error"] = str(e)
            video["finished_at"] = datetime.now().isoformat(timespec="seconds")
            video["elapsed_seconds"] = int(time.time() - start_time)


# -------------------------------------------------------------------
# Routes
# -------------------------------------------------------------------
@app.route("/", methods=["GET"])
def index():
    with BATCH_LOCK:
        batch_snapshot = BATCH.copy() if BATCH else None
        if batch_snapshot:
            videos_copy = []
            for v in BATCH["videos"]:
                vc = v.copy()
                vc["segment_display"] = format_segment_label(vc)
                videos_copy.append(vc)
            batch_snapshot["videos"] = videos_copy
    return render_template(
        "index.html",
        languages=SUPPORTED_LANGUAGES,
        max_links=MAX_LINKS_PER_BATCH,
        batch=batch_snapshot,
        downloads_info=get_downloads_info(),
        cleanup_msg=request.args.get("cleanup_msg"),
    )


@app.route("/start", methods=["POST"])
def start():
    global BATCH

    # Block starting a new batch while one is running.
    with BATCH_LOCK:
        if BATCH and BATCH["status"] == "running":
            return "A batch is already running. Wait for it to finish.", 400

    links_raw = request.form.get("links", "").strip()  # legacy, no longer used; kept for safety
    source_lang = request.form.get("source_language", "").strip()
    translate_flag = request.form.get("translate") == "1"
    target_lang = request.form.get("target_language", "").strip()

    if source_lang not in SUPPORTED_LANGUAGES:
        return f"Invalid source language: {source_lang}", 400

    if translate_flag:
        if target_lang not in SUPPORTED_LANGUAGES:
            return f"Translation is enabled but target language is invalid: {target_lang}", 400
        if target_lang == source_lang:
            return (
                f"Translation is enabled but source and target language are both '{source_lang}'. "
                f"Pick a different target language or untick the Translate checkbox.",
                400,
            )
    else:
        # When translation is off, ignore the dropdown value entirely.
        target_lang = ""

    # Prompt is now hardcoded in transcribe.DEFAULT_TRANSCRIPTION_PROMPT.
    # No longer accepted from the form. The translation variant is selected
    # downstream based on whether target_lang is non-empty.
    prompt = ""

    # Read up to MAX_LINKS_PER_BATCH separate link fields and matching title
    # and context fields. Empty rows are skipped so the user can submit 1, 2, or 3 videos.
    submissions = []  # list of (link, title, context, start_sec, end_sec) tuples in submission order
    for slot in range(1, MAX_LINKS_PER_BATCH + 1):
        link = request.form.get(f"link_{slot}", "").strip()
        title = request.form.get(f"title_{slot}", "").strip()
        context = request.form.get(f"context_{slot}", "").strip()
        start_str = request.form.get(f"start_{slot}", "").strip()
        end_str = request.form.get(f"end_{slot}", "").strip()
        if not link and not title and not context and not start_str and not end_str:
            continue  # empty row, skip
        if not link:
            return f"Row {slot} has data but no Google Drive link.", 400
        if not extract_drive_id(link):
            return f"Could not parse a Google Drive file ID from row {slot}: {link}", 400
        # Defensive cap on context length matches the textarea maxlength.
        if len(context) > 2000:
            return f"Row {slot} video context is too long ({len(context)} chars); max 2000.", 400
        # Parse optional time range. Either or both may be blank.
        try:
            start_sec = parse_mmss(start_str)
            end_sec = parse_mmss(end_str)
        except ValueError as e:
            return f"Row {slot} time range: {e}", 400
        if start_sec is not None and end_sec is not None and end_sec <= start_sec:
            return (
                f"Row {slot} time range: end time ({end_str}) must be after "
                f"start time ({start_str}).",
                400,
            )
        submissions.append((link, title, context, start_sec, end_sec))

    if not submissions:
        return "Please paste at least one Google Drive link.", 400

    batch_id = uuid.uuid4().hex[:8]
    videos = []
    for i, (link, title, context, start_sec, end_sec) in enumerate(submissions, start=1):
        videos.append({
            "index": i,
            "link": link,
            "title": title,  # may be empty string; filename code falls back to videoN
            "context": context,  # may be empty string
            "segment_start": start_sec,  # int seconds or None
            "segment_end": end_sec,  # int seconds or None
            "status": "queued",
            "step": "Queued",
            "elapsed_seconds": 0,
            "started_at": None,
            "finished_at": None,
            "error": None,
            "transcript_file": None,
            "transcript_xlsx_file": None,
            "translation_file": None,
            "log": [],
        })

    with BATCH_LOCK:
        BATCH = {
            "id": batch_id,
            "status": "running",
            "source_language": source_lang,
            "target_language": target_lang,  # empty string when translation is off
            "prompt": prompt,
            "started_at": datetime.now().isoformat(timespec="seconds"),
            "finished_at": None,
            "videos": videos,
        }

    worker = threading.Thread(target=run_batch, args=(batch_id,), daemon=True)
    worker.start()

    return redirect(url_for("index"))


@app.route("/status")
def status():
    """Return the current batch state as JSON. Used by the auto-refresh in the UI."""
    with BATCH_LOCK:
        if not BATCH:
            return jsonify({"batch": None})
        snapshot = BATCH.copy()
        snapshot["videos"] = [v.copy() for v in BATCH["videos"]]
    return jsonify({"batch": snapshot})


@app.route("/retry/<int:video_index>", methods=["POST"])
def retry(video_index):
    with BATCH_LOCK:
        if not BATCH:
            return "No active batch.", 400
        # Don't allow retry while the batch is actively processing another video.
        any_running = any(v["status"] == "running" for v in BATCH["videos"])
        if any_running:
            return "Cannot retry while another video is running.", 400
        video = next((v for v in BATCH["videos"] if v["index"] == video_index), None)
        if not video:
            return "Video not found.", 404
        if video["status"] != "failed":
            return "Only failed videos can be retried.", 400
        source_lang = BATCH["source_language"]
        target_lang = BATCH.get("target_language", "")
        prompt = BATCH["prompt"]
        batch_id = BATCH["id"]

    worker = threading.Thread(
        target=retry_single_video,
        args=(video, source_lang, target_lang, prompt, batch_id),
        daemon=True,
    )
    worker.start()
    return redirect(url_for("index"))


@app.route("/new", methods=["POST"])
def new_batch():
    """Clear the current batch (only allowed when not running)."""
    global BATCH
    with BATCH_LOCK:
        if BATCH and BATCH["status"] == "running":
            return "Cannot clear while a batch is running.", 400
        BATCH = None
    return redirect(url_for("index"))


@app.route("/cleanup", methods=["POST"])
def cleanup():
    """
    Delete files in the downloads/ folder.

    Form param 'force' = '1' deletes everything (including files reserved for
    retries of currently-failed videos). Otherwise we do a safe cleanup that
    preserves retry capability.

    Blocked while a batch is actively running (we could be yanking a file
    that's mid-transcription).
    """
    with BATCH_LOCK:
        if BATCH and BATCH["status"] == "running":
            return "Cannot clean up while a batch is running.", 400

    force = request.form.get("force") == "1"
    result = cleanup_downloads(force=force)

    # Build a short status message for the UI.
    if result["deleted_count"] == 0 and result["skipped_count"] == 0:
        msg = "Downloads folder was already empty."
    else:
        parts = [
            f"Deleted {result['deleted_count']} file(s), freed {result['deleted_mb']} MB."
        ]
        if result["skipped_count"] > 0:
            parts.append(
                f"Skipped {result['skipped_count']} file(s) reserved for retry "
                f"(use Force clean to remove)."
            )
        if result["errors"]:
            parts.append(f"Errors: {'; '.join(result['errors'][:3])}")
        msg = " ".join(parts)

    return redirect(url_for("index", cleanup_msg=msg))


@app.route("/downloads_info")
def downloads_info():
    """Return current downloads/ folder size as JSON."""
    return jsonify(get_downloads_info())


@app.route("/outputs/<path:filename>")
def download_output(filename):
    # Basic safety: only allow plain filenames, no traversal.
    if "/" in filename or ".." in filename:
        abort(404)
    return send_from_directory(OUTPUTS_DIR, filename, as_attachment=True)


# -------------------------------------------------------------------
# Entry point
# -------------------------------------------------------------------
def _run_cli_cleanup():
    """Handle `python app.py --clean-downloads` for use from the terminal."""
    info_before = get_downloads_info()
    print(
        f"Downloads folder contains {info_before['file_count']} file(s), "
        f"{info_before['total_mb']} MB total."
    )
    if info_before["file_count"] == 0:
        print("Nothing to clean up.")
        return 0
    # From the CLI, we force-delete everything. The user is explicitly asking.
    result = cleanup_downloads(force=True)
    print(
        f"Deleted {result['deleted_count']} file(s), "
        f"freed {result['deleted_mb']} MB."
    )
    if result["errors"]:
        print("Errors:")
        for err in result["errors"]:
            print(f"  - {err}")
        return 1
    return 0


# -------------------------------------------------------------------
# Transliteration page
# -------------------------------------------------------------------
ALLOWED_TRANSLITERATE_EXTS = {".srt", ".json"}


def _safe_input_filename(original: str) -> str:
    """Sanitise an uploaded filename to a safe basename. Strips paths and
    forbids unusual characters; preserves the extension."""
    name = Path(original).name  # strip any path components
    stem, ext = Path(name).stem, Path(name).suffix
    safe_stem = re.sub(r"[^a-zA-Z0-9_-]+", "_", stem).strip("_") or "input"
    return f"{safe_stem[:80]}{ext.lower()}"


@app.route("/transliterate", methods=["GET"])
def transliterate_form():
    return render_template(
        "transliterate.html",
        languages=transliterate.SUPPORTED_LANGUAGES,
        result=None,
        error=None,
    )


@app.route("/transliterate", methods=["POST"])
def transliterate_submit():
    uploaded = request.files.get("file")
    target_language = request.form.get("target_language", "").strip()

    # Validate inputs up front so the user sees errors fast.
    if not uploaded or not uploaded.filename:
        return _render_transliterate_error("Please choose a file to upload.")
    if target_language not in transliterate.SUPPORTED_LANGUAGES:
        return _render_transliterate_error(
            f"Invalid target language: {target_language!r}"
        )

    safe_name = _safe_input_filename(uploaded.filename)
    ext = Path(safe_name).suffix
    if ext not in ALLOWED_TRANSLITERATE_EXTS:
        return _render_transliterate_error(
            f"File must be .srt or .json (got {ext!r})."
        )

    # Save the upload to disk so transliterate.convert_file can read it.
    # We use a per-request unique name to avoid collisions across users.
    unique = uuid.uuid4().hex[:8]
    input_path = UPLOADS_DIR / f"{unique}_{safe_name}"
    uploaded.save(str(input_path))

    output_filename = f"{Path(safe_name).stem}_{safe_slug(target_language)}{ext}"
    output_path = OUTPUTS_DIR / output_filename

    # Accumulate progress lines so we can show them on the result page.
    log_lines: list[str] = []

    def log(msg: str):
        stamp = datetime.now().strftime("%H:%M:%S")
        log_lines.append(f"[{stamp}] {msg}")

    try:
        summary = transliterate.convert_file(
            input_path=input_path,
            output_path=output_path,
            target_language=target_language,
            log_callback=log,
        )
    except transliterate.ConversionError as e:
        return render_template(
            "transliterate.html",
            languages=transliterate.SUPPORTED_LANGUAGES,
            result=None,
            error=f"Conversion failed: {e}",
            log_lines=log_lines,
        )
    except Exception as e:
        # Anything else (parse error, value error, network error after retries)
        return render_template(
            "transliterate.html",
            languages=transliterate.SUPPORTED_LANGUAGES,
            result=None,
            error=f"Error: {e}",
            log_lines=log_lines,
        )
    finally:
        # Always delete the uploaded input file; we don't keep user uploads.
        try:
            input_path.unlink(missing_ok=True)
        except Exception:
            pass

    return render_template(
        "transliterate.html",
        languages=transliterate.SUPPORTED_LANGUAGES,
        result={
            "output_filename": output_filename,
            "target_language": target_language,
            "summary": summary,
        },
        error=None,
        log_lines=log_lines,
    )


def _render_transliterate_error(message: str):
    return render_template(
        "transliterate.html",
        languages=transliterate.SUPPORTED_LANGUAGES,
        result=None,
        error=message,
    )


@app.errorhandler(413)
def _request_entity_too_large(_e):
    """Friendly error for uploads above MAX_CONTENT_LENGTH."""
    return _render_transliterate_error(
        "Upload is too large. The maximum is 5 MB."
    ), 413


if __name__ == "__main__":
    # CLI mode: python app.py --clean-downloads
    if len(sys.argv) > 1 and sys.argv[1] == "--clean-downloads":
        sys.exit(_run_cli_cleanup())

    if not os.environ.get("GEMINI_API_KEY"):
        print("WARNING: GEMINI_API_KEY is not set. Create a .env file with:")
        print("    GEMINI_API_KEY=your-key-here")
    # threaded=True is essential so the /status endpoint stays responsive
    # while the background worker is running.
    app.run(host="127.0.0.1", port=5000, debug=False, threaded=True)
